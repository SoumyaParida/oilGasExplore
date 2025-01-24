import streamlit as st
import requests
import pandas as pd
import xlrd
from datetime import datetime
import csv
import os
import traceback as tb
from io import BytesIO
from openpyxl import load_workbook
from datetime import datetime
from dateutil.relativedelta import relativedelta
import lxml

# Define the URLs for the select options
URL_OPTIONS = {
    "North Dakota Oil and Gas": "https://www.dmr.nd.gov/oilgas/mpr/",
    "Texas Railroad Commission": "http://webapps.rrc.state.tx.us/PDQ/generalReportAction.do",
}

def pull_files_and_create_master_df(base_url, dates):
    """
    Function to fetch and process data from a given URL for specified dates.
    """
    master_df = pd.DataFrame()
    try:
        for date in dates:
            file_url = f"{base_url}{date}.xlsx"  # Assuming files are .xlsx
            response = requests.get(file_url)
            
            # Check if the file is .xlsx (openpyxl) or .xls (xlrd)
            if file_url.endswith(".xlsx"):
                # Use openpyxl for .xlsx files
                workbook = load_workbook(filename=BytesIO(response.content))
                sheet = workbook.active
                rows = list(sheet.iter_rows(values_only=True))
                df = pd.DataFrame(rows[1:], columns=rows[0])  # First row as header
            elif file_url.endswith(".xls"):
                # Use xlrd for .xls files
                workbook = xlrd.open_workbook(file_contents=response.content)
                worksheet = workbook.sheet_by_index(0)
                rows = [worksheet.row_values(row_idx) for row_idx in range(worksheet.nrows)]
                df = pd.DataFrame(rows[1:], columns=rows[0])  # First row as header
            else:
                raise ValueError(f"Unsupported file format for URL: {file_url}")
            
            # Append the data to the master DataFrame
            master_df = pd.concat([master_df, df], ignore_index=True)
    except Exception as e:
        st.error(f"Error fetching data: {str(e)}")
        tb.print_exc()
    return master_df

def texas_data(startMonth, startYear, endMonth, endYear, county_code):
    # getting fresh JSESSIONID
    s = requests.Session()
    headers = {
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
        'Accept-Language': 'en-US,en;q=0.9',
        'Connection': 'keep-alive',
        # 'Cookie': '_ga=GA1.3.1330622583.1681476442; _gid=GA1.3.910762609.1681476442; JSESSIONID=4V2AH2hx4hTPKo_zr376z59Pc6XEN01oWt99YYR8kffRCiQVNTfn!991815829',
        'Referer': 'http://webapps.rrc.state.tx.us/PDQ/home.do',
        'Upgrade-Insecure-Requests': '1',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/112.0.0.0 Safari/537.36 Edg/112.0.1722.39',
    }

    response = s.get(
        'http://webapps.rrc.state.tx.us/PDQ/generalReportAction.do',
        headers=headers,
        verify=False,
    )
    suggest = s.get('http://webapps.rrc.state.tx.us/PDQ/generalReportAction.do')
    JSESSIONID = s.cookies.get("JSESSIONID")

    print("JSESSIONID", JSESSIONID)
    print("cookies", s.cookies)

    cookies = {
        '_ga': 'GA1.3.1330622583.1681476442',
        '_gid': 'GA1.3.910762609.1681476442',
        'JSESSIONID': JSESSIONID,  ## get this fresh session id
    }

    headers = {
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
        'Accept-Language': 'en-US,en;q=0.9',
        'Cache-Control': 'max-age=0',
        'Connection': 'keep-alive',
        'Content-Type': 'application/x-www-form-urlencoded',
        # 'Cookie': '_ga=GA1.3.1330622583.1681476442; _gid=GA1.3.910762609.1681476442; JSESSIONID=VCl_znWA5_hmBrn90Y--1XPv5qSwr2fb_Rx2qXDvU60loP3GE91U!991815829',
        'Origin': 'http://webapps.rrc.state.tx.us',
        'Referer': 'http://webapps.rrc.state.tx.us/PDQ/generalReportAction.do',
        'Upgrade-Insecure-Requests': '1',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/112.0.0.0 Safari/537.36 Edg/112.0.1722.39',
    }

    data = {
        'viewType': 'Operator',
        'startMonth': startMonth,
        'startYear': startYear,
        'endMonth': endMonth,
        'endYear': endYear,
        'wellType': 'Both',
        'district': 'None Selected',
        'onShoreCounty': county_code, ## 001 for Anderson 003 for Andrews
        'offShoreArea': 'None Selected',
        'fieldNameData': 'None Selected',
        'fieldRegion': 'None Selected',
        'operatorNameData': 'None Selected',
        'operatorValue': 'None Selected',
        'submit': 'Submit',
    }

    params = {
        'pagesize': '10000',
    }

    verify = False

    months = ['01', '02', '03', '04', '05', '06', '07', '08', '09', '10', '11', '12']
    month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    year = ['2022', '2023']

    response = requests.post(
        'http://webapps.rrc.state.tx.us/PDQ/mainReportAction.do',
        cookies=cookies,
        headers=headers,
        data=data,
        verify=False,
    )

    print(response.status_code)

    # Use lxml for parsing HTML tables
    df = pd.read_html(response.text, flavor='lxml')

    #print(df[12])
    return df[12]

# Streamlit UI
st.title("Energy Data Fetcher")

# Select website
selected_site = st.selectbox("Choose a website to fetch data:", list(URL_OPTIONS.keys()))

# Date range selection
start_date = st.date_input(
    "Start Date",
    value=datetime(2020, 1, 1),
    min_value=datetime(1990, 1, 1),
    max_value=datetime.now()
)
end_date = st.date_input(
    "End Date",
    value=datetime.now(),
    min_value=start_date,
    max_value=datetime.now()
)

# Extract start and end years and months
start_year = start_date.year
start_month = start_date.month
end_year = end_date.year
end_month = end_date.month

# Generate the list of dates
dates = []
current_date = datetime(start_year, start_month, 1)

while current_date <= datetime(end_year, end_month, 1):
    dates.append(current_date.strftime('%Y_%m'))  # Format as 'YYYY_MM'
    current_date += relativedelta(months=1)  # Increment by 1 month

# For Texas, additional inputs are required
if selected_site == "Texas Railroad Commission":
    st.subheader("Additional Inputs for Texas")
    county_code = st.text_input("Enter Onshore County Code (e.g., 001 for Anderson):")

if st.button("Fetch Data"):
    if selected_site == "North Dakota Oil and Gas":
        base_url = URL_OPTIONS[selected_site]
        master_df = pull_files_and_create_master_df(base_url, dates)
        st.write("Data fetched successfully!")
        st.write(master_df)
    elif selected_site == "Texas Railroad Commission":
        texas_df = texas_data(str(start_month), str(start_year), str(end_month), str(end_year), county_code)
        st.write(texas_df)